import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from datetime import datetime
import pytz
import os

# --- Konstanta Global ---
KEPALA_SEKOLAH = "Alina Fiftiyani Nurjannah, M.Pd."
NIP_KEPSEK = "19800105 200903 2 006"

# --- Daftar Mata Pelajaran Umum (untuk selectbox) ---
COMMON_SUBJECTS = [
    "Matematika", "Bahasa Indonesia", "IPA",
    "IPS", "Bahasa Inggris", "Pendidikan Agama Islam",
    "PP", "Seni Budaya",
    "PJOK", "Informatika",
    "Prakarya", "Bahasa Jawa"
]

# =========================================================
# Fungsi untuk membuat dokumen Excel "Daftar Hadir Siswa"
# =========================================================
def generate_excel_absensi_panjang(dataframe, mapel, semester, kelas, tahun_pelajaran, guru, nip_guru):
    wb = Workbook()
    ws = wb.active
    ws.title = f"Daftar Hadir {kelas}"

    # --- Pengaturan Halaman (Legal, Portrait) ---
    ws.page_setup.paperSize = ws.PAPERSIZE_LEGAL # Paper size Legal
    # Margins in inches (1 inch = 25.4 mm)
    ws.page_margins = PageMargins(
        left=18/25.4, # Kiri 18mm
        right=0.8/25.4, # Kanan 0.8mm
        top=19/25.4, # Atas 19mm
        bottom=24/25.4 # Bawah 24mm
    )
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT # Portrait
    ws.print_options.horizontalCentered = True # Rata Halaman Horizontal
    ws.print_options.verticalCentered = False # Rata Halaman Vertikal dinonaktifkan

    # --- Styles ---
    font_bold_12 = Font(name='Times New Roman', size=12, bold=True)
    font_normal_12 = Font(name='Times New Roman', size=12)
    font_bold_9 = Font(name='Times New Roman', size=9, bold=True)
    font_normal_9 = Font(name='Times New Roman', size=9)
    font_normal_7 = Font(name='Times New Roman', size=7) # <--- BARU: Font ukuran 7

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=False)
    right_align = Alignment(horizontal='right', vertical='center', wrap_text=False)

    # --- Header Informasi (Baris atas) ---
    current_row = 1

    # Define constants for column ranges based on new requirement (W-AA for meeting columns)
    COL_NO_START = 1 # A
    COL_NO_END = 2 # B
    COL_NAMA_START = 3 # C
    COL_NAMA_END = 4 # D
    COL_PERTEMUAN_START = 5 # E
    # If W-AA (23-27) are added, and assuming 20 total meeting columns (15 + 5)
    COL_PERTEMUAN_END = 24 # X (E to X is 20 columns: 5 + 20 - 1 = 24)
    COL_JUMLAH_START = 25 # Y
    COL_JUMLAH_END = 27 # AA (Y to AA is 3 columns: 25 + 3 - 1 = 27)

    TOTAL_EFFECTIVE_COLS = 27 # A (1) to AA (27)

    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=TOTAL_EFFECTIVE_COLS)
    cell_title = ws[f'A{current_row}']
    cell_title.value = "DAFTAR HADIR SISWA"
    cell_title.font = font_bold_12
    cell_title.alignment = center_align

    current_row += 1
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=TOTAL_EFFECTIVE_COLS)
    cell_subtitle = ws[f'A{current_row}']
    cell_subtitle.value = f"SMP NEGERI 2 BANGUNTAPAN TAHUN PELAJARAN {tahun_pelajaran}"
    cell_subtitle.font = font_bold_12
    cell_subtitle.alignment = center_align

    current_row += 2 # Spasi setelah judul

    # Information table (Mata Pelajaran, Kelas, Semester, Tahun Pelajaran)
    ws[f'A{current_row}'].value = "Mata Pelajaran"
    ws[f'D{current_row}'].value = f": {mapel}"
    ws[f'K{current_row}'].value = "Semester"
    ws[f'T{current_row}'].value = f": {semester}"

    for col_letter in ['A', 'D', 'K', 'T']:
        ws[f'{col_letter}{current_row}'].font = font_normal_12
        ws[f'{col_letter}{current_row}'].alignment = left_align

    current_row += 1
    ws[f'A{current_row}'].value = "Kelas"
    ws[f'D{current_row}'].value = f": {kelas}"
    ws[f'K{current_row}'].value = "Tahun Pelajaran"
    ws[f'T{current_row}'].value = f": {tahun_pelajaran}"

    for col_letter in ['A', 'D', 'K', 'T']:
        ws[f'{col_letter}{current_row}'].font = font_normal_12
        ws[f'{col_letter}{current_row}'].alignment = left_align

    current_row += 2 # Spasi sebelum tabel utama

    # --- Tabel Utama Header ---
    header_start_row = current_row

    # --- Header "No." ---
    ws.merge_cells(start_row=header_start_row, start_column=COL_NO_START, end_row=header_start_row + 2, end_column=COL_NO_END)
    cell_no_header = ws.cell(row=header_start_row, column=COL_NO_START)
    cell_no_header.value = "No."
    cell_no_header.font = font_bold_12
    cell_no_header.alignment = center_align

    # --- Header "Nama" ---
    ws.merge_cells(start_row=header_start_row, start_column=COL_NAMA_START, end_row=header_start_row + 2, end_column=COL_NAMA_END)
    cell_nama_header = ws.cell(row=header_start_row, column=COL_NAMA_START)
    cell_nama_header.value = "Nama"
    cell_nama_header.font = font_bold_12
    cell_nama_header.alignment = center_align

    # --- Header "Pertemuan ke... Tanggal..." ---
    ws.merge_cells(start_row=header_start_row, start_column=COL_PERTEMUAN_START, end_row=header_start_row, end_column=COL_PERTEMUAN_END)
    cell_pertemuan_header = ws.cell(row=header_start_row, column=COL_PERTEMUAN_START)
    cell_pertemuan_header.value = "Pertemuan ke... Tanggal..."
    cell_pertemuan_header.font = font_bold_12
    cell_pertemuan_header.alignment = center_align

    # --- Header "Jumlah" ---
    ws.merge_cells(start_row=header_start_row, start_column=COL_JUMLAH_START, end_row=header_start_row, end_column=COL_JUMLAH_END)
    cell_jumlah_header = ws.cell(row=header_start_row, column=COL_JUMLAH_START)
    cell_jumlah_header.value = "Jumlah"
    cell_jumlah_header.font = font_bold_9
    cell_jumlah_header.alignment = center_align

    # --- Sub-headers (Row 2 and 3 relative to header_start_row) ---
    # Pertemuan 1-20 (Row 2, E-X)
    for i in range(20): # Changed from 15 to 20 meetings
        col_idx = COL_PERTEMUAN_START + i
        cell = ws.cell(row=header_start_row + 1, column=col_idx) # Row 2
        cell.value = str(i + 1) # <--- KEMBALI KE NILAI ASLI (1, 2, 3...)
        cell.font = font_normal_7 # <--- PERUBAHAN DI SINI: Menggunakan font_normal_7
        cell.alignment = center_align

        # Third row, empty cells for E-X
        cell_empty = ws.cell(row=header_start_row + 2, column=col_idx) # Row 3
        cell_empty.value = ""
        cell_empty.font = font_normal_7 # <--- PERUBAHAN DI SINI: Menggunakan font_normal_7
        cell_empty.alignment = center_align

    # A, I, S headers (Row 2 merged with Row 3, within Jumlah Y-AA)
    # A (Absen)
    ws.merge_cells(start_row=header_start_row + 1, start_column=COL_JUMLAH_START, end_row=header_start_row + 2, end_column=COL_JUMLAH_START)
    cell_a_header = ws.cell(row=header_start_row + 1, column=COL_JUMLAH_START)
    cell_a_header.value = "A"
    cell_a_header.font = font_bold_9
    cell_a_header.alignment = center_align

    # I (Izin)
    ws.merge_cells(start_row=header_start_row + 1, start_column=COL_JUMLAH_START + 1, end_row=header_start_row + 2, end_column=COL_JUMLAH_START + 1)
    cell_i_header = ws.cell(row=header_start_row + 1, column=COL_JUMLAH_START + 1)
    cell_i_header.value = "I"
    cell_i_header.font = font_bold_9
    cell_i_header.alignment = center_align

    # S (Sakit)
    ws.merge_cells(start_row=header_start_row + 1, start_column=COL_JUMLAH_START + 2, end_row=header_start_row + 2, end_column=COL_JUMLAH_START + 2)
    cell_s_header = ws.cell(row=header_start_row + 1, column=COL_JUMLAH_START + 2)
    cell_s_header.value = "S"
    cell_s_header.font = font_bold_9
    cell_s_header.alignment = center_align

    # Apply borders to all header cells (including merged ones)
    for r_idx in range(header_start_row, header_start_row + 3):
        for c_idx in range(1, TOTAL_EFFECTIVE_COLS + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.border = thin_border

    # Set header row heights (adjust as needed for aesthetics)
    ws.row_dimensions[header_start_row].height = 20
    ws.row_dimensions[header_start_row + 1].height = 20
    ws.row_dimensions[header_start_row + 2].height = 20


    # --- Atur lebar kolom ---
    ws.column_dimensions[get_column_letter(COL_NO_START)].width = 2.57
    ws.column_dimensions[get_column_letter(COL_NO_END)].width = 3.43

    ws.column_dimensions[get_column_letter(COL_NAMA_START)].width = 8.14
    ws.column_dimensions[get_column_letter(COL_NAMA_END)].width = 26.86

    for i in range(COL_PERTEMUAN_START, COL_PERTEMUAN_END + 1): # E to X (20 columns)
        ws.column_dimensions[get_column_letter(i)].width = 2.3

    for i in range(COL_JUMLAH_START, COL_JUMLAH_END + 1): # Y to AA (3 columns)
        ws.column_dimensions[get_column_letter(i)].width = 2.7

    current_row += 3 # Move to the row where data starts after 3 header rows

    # --- Isi Data Siswa (dengan border di semua kolom) ---
    for i, row_data in enumerate(dataframe.itertuples(), 1):
        data_row_idx = current_row + i - 1

        ws.merge_cells(start_row=data_row_idx, start_column=COL_NO_START, end_row=data_row_idx, end_column=COL_NO_END)
        ws.merge_cells(start_row=data_row_idx, start_column=COL_NAMA_START, end_row=data_row_idx, end_column=COL_NAMA_END)

        ws.cell(row=data_row_idx, column=COL_NO_START, value=i).font = font_normal_12
        ws.cell(row=data_row_idx, column=COL_NO_START).alignment = center_align

        ws.cell(row=data_row_idx, column=COL_NAMA_START, value=getattr(row_data, "Nama")).font = font_normal_12
        ws.cell(row=data_row_idx, column=COL_NAMA_START).alignment = left_align

        # Apply border to all cells in the row, AFTER merges
        for c_idx in range(1, TOTAL_EFFECTIVE_COLS + 1): # Iterate from column A (1) to AA (27)
            cell = ws.cell(row=data_row_idx, column=c_idx)
            cell.border = thin_border
            # Values for columns E-AA are empty by default, so no need to set value again for meeting/absentee columns

        ws.row_dimensions[data_row_idx].height = 20

    current_row += len(dataframe) + 1 # Pindah ke bawah tabel, 1 baris spasi

    # --- Bagian Tanda Tangan ---
    ws.cell(row=current_row, column=3, value="Mengetahui").font = font_normal_12
    ws.cell(row=current_row, column=3).alignment = left_align
    current_row += 1
    ws.cell(row=current_row, column=3, value="Kepala Sekolah").font = font_normal_12
    ws.cell(row=current_row, column=3).alignment = left_align

    tz_jakarta = pytz.timezone('Asia/Jakarta')
    now = datetime.now(tz_jakarta)

    # Format Tanggal: Titik-titik untuk Tanggal dan Bulan, Tahun sesuai tahun berlaku
    tanggal_formatted = f"Bantul, .................................... {now.year}"

    ws.cell(row=current_row - 1, column=12, value=tanggal_formatted).font = font_normal_12
    ws.cell(row=current_row - 1, column=12).alignment = left_align
    ws.cell(row=current_row, column=12, value=f"Guru {mapel}").font = font_normal_12
    ws.cell(row=current_row, column=12).alignment = left_align

    current_row += 3 # Jarak untuk tanda tangan (3 baris kosong)

    ws.cell(row=current_row, column=3, value=KEPALA_SEKOLAH).font = font_normal_12
    ws.cell(row=current_row, column=3).alignment = left_align
    ws.cell(row=current_row, column=12, value=guru).font = font_normal_12
    ws.cell(row=current_row, column=12).alignment = left_align

    current_row += 1
    ws.cell(row=current_row, column=3, value=f"NIP. {NIP_KEPSEK}").font = font_normal_12
    ws.cell(row=current_row, column=3).alignment = left_align
    ws.cell(row=current_row, column=12, value=f"NIP. {nip_guru}").font = font_normal_12
    ws.cell(row=current_row, column=12).alignment = left_align

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# =========================================================
# Fungsi untuk membuat dokumen Excel "Daftar Siswa"
# =========================================================
def generate_excel_daftar_siswa(dataframe, kelas, semester, tahun_pelajaran, nama_wali_kelas, nip_wali_kelas):
    wb = Workbook()
    ws = wb.active
    ws.title = f"Daftar Siswa {kelas}"

    # --- Pengaturan Halaman (Legal, Portrait) ---
    ws.page_setup.paperSize = ws.PAPERSIZE_LEGAL # Paper size Legal
    # Margins in inches (1 inch = 25.4 mm)
    ws.page_margins = PageMargins(left=28/25.4, right=18/25.4, top=24/25.4, bottom=19/25.4)
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT # Portrait
    ws.print_options.horizontalCentered = True # Rata Halaman Horizontal
    ws.print_options.verticalCentered = False # Rata Halaman Vertikal dinonaktifkan

    # --- Styles ---
    font_bold_12 = Font(name='Times New Roman', size=12, bold=True)
    font_normal_12 = Font(name='Times New Roman', size=12)
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=False)

    # --- Header Dokumen Excel ---
    current_row = 1
    # Merge cells for main title (spans 7 columns: A-G)
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
    cell_title = ws[f'A{current_row}']
    cell_title.value = "DAFTAR SISWA"
    cell_title.font = font_bold_12
    cell_title.alignment = center_align

    current_row += 1
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
    cell_subtitle = ws[f'A{current_row}']
    cell_subtitle.value = f"SMP NEGERI 2 BANGUNTAPAN TAHUN PELAJARAN {tahun_pelajaran}"
    cell_subtitle.font = font_bold_12
    cell_subtitle.alignment = center_align

    current_row += 2 # Spasi setelah judul

    # Information table
    # Kelas dan Wali Kelas di kolom A, isi kelas dan isi Nama Wali kelas di kolom D
    ws[f'A{current_row}'].value = "Kelas"
    ws[f'D{current_row}'].value = f": {kelas}"
    # Semester dan Tahun Pelajaran di kolom F, isi Semester dan isi Tahun Pelajaran di kolom G
    ws[f'F{current_row}'].value = "Semester"
    ws[f'G{current_row}'].value = f": {semester}"

    for col_letter in ['A', 'D', 'F', 'G']:
        ws[f'{col_letter}{current_row}'].font = font_normal_12
        ws[f'{col_letter}{current_row}'].alignment = left_align # All info cells left aligned and vertically centered

    current_row += 1
    ws[f'A{current_row}'].value = "Wali Kelas" # Changed from "Nama Wali Kelas"
    ws[f'D{current_row}'].value = f": {nama_wali_kelas}"
    ws[f'F{current_row}'].value = "Tahun Pelajaran"
    ws[f'G{current_row}'].value = f": {tahun_pelajaran}"

    for col_letter in ['A', 'D', 'F', 'G']:
        ws[f'{col_letter}{current_row}'].font = font_normal_12
        ws[f'{col_letter}{current_row}'].alignment = left_align # All info cells left aligned and vertically centered

    current_row += 2 # Spasi sebelum tabel siswa

    # --- Tabel Siswa Header ---
    header_start_row = current_row

    # Define new column constants for the table based on swapped NIS and Nama Siswa
    COL_NO_START = 1
    COL_NO_END = 2 # Merged with COL_NO_START (A:B)
    COL_NIS_START = 3 # NIS is now column C
    COL_NIS_END = 4 # Merged with COL_NIS_START (C:D)
    COL_NAMA_START = 5 # Nama Siswa is now column E
    COL_NAMA_END = 6 # Merged with COL_NAMA_START (E:F)
    COL_JK = 7 # Jenis Kelamin is now column G
    TOTAL_TABLE_COLS = 7 # Total columns in this table

    # Merge headers for No.
    ws.merge_cells(start_row=header_start_row, start_column=COL_NO_START, end_row=header_start_row, end_column=COL_NO_END)
    ws.cell(row=header_start_row, column=COL_NO_START, value="No.").font = font_bold_12
    ws.cell(row=header_start_row, column=COL_NO_START).alignment = center_align

    # Merge headers for NIS
    ws.merge_cells(start_row=header_start_row, start_column=COL_NIS_START, end_row=header_start_row, end_column=COL_NIS_END)
    ws.cell(row=header_start_row, column=COL_NIS_START, value="NIS").font = font_bold_12
    ws.cell(row=header_start_row, column=COL_NIS_START).alignment = center_align

    # Merge headers for Nama Siswa
    ws.merge_cells(start_row=header_start_row, start_column=COL_NAMA_START, end_row=header_start_row, end_column=COL_NAMA_END)
    ws.cell(row=header_start_row, column=COL_NAMA_START, value="Nama Siswa").font = font_bold_12
    ws.cell(row=header_start_row, column=COL_NAMA_START).alignment = left_align # Nama Siswa header should be left aligned

    ws.cell(row=header_start_row, column=COL_JK, value="JK").font = font_bold_12
    ws.cell(row=header_start_row, column=COL_JK).alignment = center_align

    # Set header row height
    ws.row_dimensions[header_start_row].height = 21

    # Apply borders and vertical center alignment to header cells
    for c_idx in range(1, TOTAL_TABLE_COLS + 1):
        cell = ws.cell(row=header_start_row, column=c_idx)
        cell.border = thin_border
        # Alignment already set for each cell above, no need to re-apply

    current_row += 1 # Pindah ke baris untuk entri data

    # --- Isi Data Siswa ---
    for i, row_data in enumerate(dataframe.itertuples(), 1):
        data_row_idx = current_row + i - 1

        # Merge cells for No. and NIS for each data row
        ws.merge_cells(start_row=data_row_idx, start_column=COL_NO_START, end_row=data_row_idx, end_column=COL_NO_END)
        ws.merge_cells(start_row=data_row_idx, start_column=COL_NIS_START, end_row=data_row_idx, end_column=COL_NIS_END)
        # Merge cells for Nama Siswa for each data row
        ws.merge_cells(start_row=data_row_idx, start_column=COL_NAMA_START, end_row=data_row_idx, end_column=COL_NAMA_END)

        ws.cell(row=data_row_idx, column=COL_NO_START, value=i).font = font_normal_12
        ws.cell(row=data_row_idx, column=COL_NIS_START, value=str(getattr(row_data, "NIS"))).font = font_normal_12
        ws.cell(row=data_row_idx, column=COL_NAMA_START, value=getattr(row_data, "Nama")).font = font_normal_12
        ws.cell(row=data_row_idx, column=COL_JK, value=getattr(row_data, "Jenis_Kelamin")).font = font_normal_12

        # Set row height for data rows
        ws.row_dimensions[data_row_idx].height = 17.25

        # Apply borders and alignment to data cells
        for col_idx in range(1, TOTAL_TABLE_COLS + 1):
            cell = ws.cell(row=data_row_idx, column=col_idx)
            cell.border = thin_border
            if col_idx == COL_NAMA_START: # Nama Siswa
                cell.alignment = left_align
            else:
                cell.alignment = center_align

    # Atur lebar kolom
    ws.column_dimensions[get_column_letter(COL_NO_START)].width = 2.29 # No (merged)
    ws.column_dimensions[get_column_letter(COL_NO_END)].width = 2.14 # Hide the merged part
    ws.column_dimensions[get_column_letter(COL_NIS_START)].width = 6.14 # NIS (merged)
    ws.column_dimensions[get_column_letter(COL_NIS_END)].width = 1.71 # Hide the merged part
    ws.column_dimensions[get_column_letter(COL_NAMA_START)].width = 22.71 # Nama Siswa (merged)
    ws.column_dimensions[get_column_letter(COL_NAMA_END)].width = 27 # Hide the merged part
    ws.column_dimensions[get_column_letter(COL_JK)].width = 7.86 # Jenis Kelamin

    current_row += len(dataframe) + 1 # Pindah ke bawah tabel, 1 baris spasi

    # --- Bagian Keterangan ---
    if 'Jenis_Kelamin' in dataframe.columns:
        jumlah_L = dataframe[dataframe['Jenis_Kelamin'].astype(str).str.upper() == 'L'].shape[0]
        jumlah_P = dataframe[dataframe['Jenis_Kelamin'].astype(str).str.upper() == 'P'].shape[0]

        ws.cell(row=current_row, column=1, value="Keterangan:").font = font_bold_12
        ws.cell(row=current_row, column=1).alignment = left_align
        current_row += 1
        ws.cell(row=current_row, column=1, value=f"- Jumlah Laki-laki (L) : {jumlah_L}").font = font_normal_12
        ws.cell(row=current_row, column=1).alignment = left_align
        current_row += 1
        ws.cell(row=current_row, column=1, value=f"- Jumlah Perempuan (P) : {jumlah_P}").font = font_normal_12
        ws.cell(row=current_row, column=1).alignment = left_align
        current_row += 1

    current_row += 1 # Tambah 1 baris kosong antara Keterangan dan Mengetahui

    # --- Bagian Tanda Tangan ---
    #Mengetahui (Kepala Sekolah) di kolom A
    ws.cell(row=current_row, column=1, value="Mengetahui").font = font_normal_12
    ws.cell(row=current_row, column=1).alignment = left_align
    ws.cell(row=current_row + 1, column=1, value="Kepala Sekolah").font = font_normal_12
    ws.cell(row=current_row + 1, column=1).alignment = left_align

    #Wali Kelas dan Bantul di kolom F (index 7)
    ws.cell(row=current_row, column=6, value=f"Bantul, ............................... {tahun_pelajaran.split('/')[0]}").font = font_normal_12
    ws.cell(row=current_row, column=6).alignment = left_align
    ws.cell(row=current_row + 1, column=6, value=f"Wali Kelas {kelas}").font = font_normal_12
    ws.cell(row=current_row + 1, column=6).alignment = left_align
    current_row += 2

    current_row += 3 # Jarak untuk tanda tangan (3 baris)
    ws.cell(row=current_row, column=1, value=KEPALA_SEKOLAH).font = font_normal_12
    ws.cell(row=current_row, column=1).alignment = left_align
    ws.cell(row=current_row, column=6, value=nama_wali_kelas).font = font_normal_12
    ws.cell(row=current_row, column=6).alignment = left_align

    current_row += 1 # Pindah ke baris berikutnya untuk NIP
    ws.cell(row=current_row, column=1, value=f"NIP. {NIP_KEPSEK}").font = font_normal_12
    ws.cell(row=current_row, column=1).alignment = left_align
    ws.cell(row=current_row, column=6, value=f"NIP. {nip_wali_kelas}").font = font_normal_12
    ws.cell(row=current_row, column=6).alignment = left_align

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# =========================================================
# Fungsi untuk membuat dokumen Excel "Form Nilai Siswa"
# =========================================================
def generate_excel_form_nilai_siswa(dataframe, mapel, semester, kelas, tahun_pelajaran, guru, nip_guru):
    wb = Workbook()
    ws = wb.active
    ws.title = f"Form Nilai {kelas}"

    # --- Pengaturan Halaman (Legal, Portrait) ---
    ws.page_setup.paperSize = ws.PAPERSIZE_LEGAL
    ws.page_margins = PageMargins(
        left=1.8/2.54, # Kiri 1.8cm
        right=1.8/2.54, # Kanan 1.8cm
        top=1.9/2.54, # Atas 1.9cm
        bottom=1.9/2.54 # Bawah 1.9cm
    )
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = False

    # --- Styles ---
    font_bold_12 = Font(name='Times New Roman', size=12, bold=True)
    font_normal_12 = Font(name='Times New Roman', size=12)
    font_bold_9 = Font(name='Times New Roman', size=9, bold=True)
    font_normal_9 = Font(name='Times New Roman', size=9)

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=False)

    # --- Header Informasi (Baris atas) ---
    current_row = 1
    # Total columns for merging the title: A to T (20 columns)
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=20)
    cell_title = ws[f'A{current_row}']
    cell_title.value = "FORM NILAI SISWA"
    cell_title.font = font_bold_12
    cell_title.alignment = center_align

    current_row += 1
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=20)
    cell_subtitle = ws[f'A{current_row}']
    cell_subtitle.value = f"SMP NEGERI 2 BANGUNTAPAN TAHUN PELAJARAN {tahun_pelajaran}"
    cell_subtitle.font = font_bold_12
    cell_subtitle.alignment = center_align

    current_row += 2 # Spasi setelah judul

    # Information table (Mata Pelajaran, Kelas, Semester, Tahun Pelajaran)
    # Mata Pelajaran dan Semester di kolom A, isi di F
    ws[f'A{current_row}'].value = "Mata Pelajaran"
    ws[f'G{current_row}'].value = f": {mapel}"
    ws[f'A{current_row}'].font = font_normal_12
    ws[f'F{current_row}'].font = font_normal_12
    ws[f'A{current_row}'].alignment = left_align
    ws[f'G{current_row}'].alignment = left_align

    # Kelas dan Nama Guru di kolom K, isi di O
    ws[f'K{current_row}'].value = "Kelas"
    ws[f'O{current_row}'].value = f": {kelas}"
    ws[f'K{current_row}'].font = font_normal_12
    ws[f'O{current_row}'].font = font_normal_12
    ws[f'K{current_row}'].alignment = left_align
    ws[f'O{current_row}'].alignment = left_align

    current_row += 1
    ws[f'A{current_row}'].value = "Semester"
    ws[f'G{current_row}'].value = f": {semester}"
    ws[f'A{current_row}'].font = font_normal_12
    ws[f'G{current_row}'].font = font_normal_12
    ws[f'A{current_row}'].alignment = left_align
    ws[f'G{current_row}'].alignment = left_align

    ws[f'K{current_row}'].value = "Nama Guru"
    ws[f'O{current_row}'].value = f": {guru}"
    ws[f'K{current_row}'].font = font_normal_12
    ws[f'O{current_row}'].font = font_normal_12
    ws[f'K{current_row}'].alignment = left_align
    ws[f'O{current_row}'].alignment = left_align

    current_row += 2 # Spasi sebelum tabel utama

    # --- Tabel Utama Header ---
    header_start_row = current_row

    # Column Constants for Form Nilai Siswa
    COL_NO_START = 1 # A
    COL_NO_END = 2 # B
    COL_NIS_START = 3 # C
    COL_NIS_END = 5 # E
    COL_NAMA_START = 6 # F
    COL_NAMA_END = 7 # G
    COL_FORMATIF_START = 8 # H
    COL_FORMATIF_END = 12 # L (H, I, J, K, L - 5 columns)
    COL_SUMATIF_START = 13 # M
    COL_SUMATIF_END = 17 # Q (M, N, O, P, Q - 5 columns)
    COL_PTS = 18 # R
    COL_SAS_SAT = 19 # S
    COL_NR = 20 # T

    TOTAL_EFFECTIVE_COLS_NILAI = COL_NR

    # Row 1 Headers
    # No.
    ws.merge_cells(start_row=header_start_row, start_column=COL_NO_START, end_row=header_start_row + 2, end_column=COL_NO_END)
    cell_no_header = ws.cell(row=header_start_row, column=COL_NO_START)
    cell_no_header.value = "No."
    cell_no_header.font = font_bold_12
    cell_no_header.alignment = center_align

    # NIS
    ws.merge_cells(start_row=header_start_row, start_column=COL_NIS_START, end_row=header_start_row + 2, end_column=COL_NIS_END)
    cell_nis_header = ws.cell(row=header_start_row, column=COL_NIS_START)
    cell_nis_header.value = "NIS"
    cell_nis_header.font = font_bold_12
    cell_nis_header.alignment = center_align

    # Nama Siswa
    ws.merge_cells(start_row=header_start_row, start_column=COL_NAMA_START, end_row=header_start_row + 2, end_column=COL_NAMA_END)
    cell_nama_header = ws.cell(row=header_start_row, column=COL_NAMA_START)
    cell_nama_header.value = "Nama Siswa"
    cell_nama_header.font = font_bold_12
    cell_nama_header.alignment = center_align

    # Formatif/Tugas
    ws.merge_cells(start_row=header_start_row, start_column=COL_FORMATIF_START, end_row=header_start_row, end_column=COL_FORMATIF_END)
    cell_formatif_header = ws.cell(row=header_start_row, column=COL_FORMATIF_START)
    cell_formatif_header.value = "Formatif / Tugas"
    cell_formatif_header.font = font_bold_9
    cell_formatif_header.alignment = center_align

    # SUMATIF LINGKUP MATERI
    ws.merge_cells(start_row=header_start_row, start_column=COL_SUMATIF_START, end_row=header_start_row, end_column=COL_SUMATIF_END)
    cell_sumatif_header = ws.cell(row=header_start_row, column=COL_SUMATIF_START)
    cell_sumatif_header.value = "Sumatif Lingkup Materi"
    cell_sumatif_header.font = font_bold_9
    cell_sumatif_header.alignment = center_align

    # PTS
    ws.merge_cells(start_row=header_start_row, start_column=COL_PTS, end_row=header_start_row + 2, end_column=COL_PTS)
    cell_pts_header = ws.cell(row=header_start_row, column=COL_PTS)
    cell_pts_header.value = "PTS"
    cell_pts_header.font = font_bold_9
    cell_pts_header.alignment = center_align

    # SAS/SAT
    ws.merge_cells(start_row=header_start_row, start_column=COL_SAS_SAT, end_row=header_start_row + 2, end_column=COL_SAS_SAT)
    cell_sas_sat_header = ws.cell(row=header_start_row, column=COL_SAS_SAT)
    cell_sas_sat_header.value = "SAS/SAT"
    cell_sas_sat_header.font = font_bold_9
    cell_sas_sat_header.alignment = center_align

    # NR
    ws.merge_cells(start_row=header_start_row, start_column=COL_NR, end_row=header_start_row + 2, end_column=COL_NR)
    cell_nr_header = ws.cell(row=header_start_row, column=COL_NR)
    cell_nr_header.value = "NR"
    cell_nr_header.font = font_bold_9
    cell_nr_header.alignment = center_align

    # Row 2 Headers (Lingkup Materi for Formatif/Tugas)
    for i in range(5): # H to L
        col_idx = COL_FORMATIF_START + i
        cell = ws.cell(row=header_start_row + 1, column=col_idx) # Row 2
        cell.value = f"TP{i+1}" # Changed from "TP1" etc. to match TP1-TP5
        cell.font = font_normal_9
        cell.alignment = center_align
        # Merge with row 3 for Formatif/Tugas LM headers
        ws.merge_cells(start_row=header_start_row + 1, start_column=col_idx, end_row=header_start_row + 2, end_column=col_idx)


    # Row 3 Headers (LM1-LM5 for SUMATIF LINGKUP MATERI)
    for i in range(5): # M to Q
        col_idx = COL_SUMATIF_START + i
        cell = ws.cell(row=header_start_row + 1, column=col_idx) # Row 3
        cell.value = f"LM{i+1}"
        cell.font = font_normal_9
        cell.alignment = center_align
        ws.merge_cells(start_row=header_start_row + 1, start_column=col_idx, end_row=header_start_row + 2, end_column=col_idx)


    # Apply borders to all header cells (including merged ones)
    for r_idx in range(header_start_row, header_start_row + 3):
        for c_idx in range(1, TOTAL_EFFECTIVE_COLS_NILAI + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.border = thin_border

    # Set header row heights (adjust as needed for aesthetics)
    ws.row_dimensions[header_start_row].height = 20
    ws.row_dimensions[header_start_row + 1].height = 20
    ws.row_dimensions[header_start_row + 2].height = 20

    # --- Atur lebar kolom ---
    ws.column_dimensions[get_column_letter(COL_NO_START)].width = 2.57
    ws.column_dimensions[get_column_letter(COL_NO_END)].width = 2.14

    ws.column_dimensions[get_column_letter(COL_NIS_START)].width = 2
    ws.column_dimensions[get_column_letter(COL_NIS_START + 1)].width = 1.14
    ws.column_dimensions[get_column_letter(COL_NIS_END)].width = 2

    ws.column_dimensions[get_column_letter(COL_NAMA_START)].width = 8.43
    ws.column_dimensions[get_column_letter(COL_NAMA_END)].width = 21.57

    for i in range(COL_FORMATIF_START, COL_FORMATIF_END + 1): # H to L
        ws.column_dimensions[get_column_letter(i)].width = 4

    for i in range(COL_SUMATIF_START, COL_SUMATIF_END + 1): # M to Q
        ws.column_dimensions[get_column_letter(i)].width = 4

    ws.column_dimensions[get_column_letter(COL_PTS)].width = 4
    ws.column_dimensions[get_column_letter(COL_SAS_SAT)].width = 5
    ws.column_dimensions[get_column_letter(COL_NR)].width = 4

    current_row += 3 # Move to the row where data starts after 3 header rows

    # --- Isi Data Siswa (dengan border di semua kolom) ---
    for i, row_data in enumerate(dataframe.itertuples(), 1):
        data_row_idx = current_row + i - 1

        # No.
        ws.merge_cells(start_row=data_row_idx, start_column=COL_NO_START, end_row=data_row_idx, end_column=COL_NO_END)
        ws.cell(row=data_row_idx, column=COL_NO_START, value=i).font = font_normal_9
        ws.cell(row=data_row_idx, column=COL_NO_START).alignment = center_align

        # NIS
        nis_value = getattr(row_data, "NIS", "") if "NIS" in dataframe.columns else ""
        ws.merge_cells(start_row=data_row_idx, start_column=COL_NIS_START, end_row=data_row_idx, end_column=COL_NIS_END)
        ws.cell(row=data_row_idx, column=COL_NIS_START, value=nis_value).font = font_normal_9
        ws.cell(row=data_row_idx, column=COL_NIS_START).alignment = center_align

        # Nama Siswa
        nama_value = getattr(row_data, "Nama", "") if "Nama" in dataframe.columns else ""
        ws.merge_cells(start_row=data_row_idx, start_column=COL_NAMA_START, end_row=data_row_idx, end_column=COL_NAMA_END)
        ws.cell(row=data_row_idx, column=COL_NAMA_START, value=nama_value).font = font_normal_9
        ws.cell(row=data_row_idx, column=COL_NAMA_START).alignment = left_align

        # Apply border to all cells in the row, AFTER merges
        for c_idx in range(1, TOTAL_EFFECTIVE_COLS_NILAI + 1): # Iterate from column A (1) to T (20)
            cell = ws.cell(row=data_row_idx, column=c_idx)
            cell.border = thin_border
            # Values for score columns are empty by default

        ws.row_dimensions[data_row_idx].height = 20

    current_row += len(dataframe) + 1 # Pindah ke bawah tabel, 1 baris spasi

    # --- Bagian Tanda Tangan ---
    # Mengetahui (Kepala Sekolah) di kolom C
    ws.cell(row=current_row, column=3, value="Mengetahui").font = font_normal_12
    ws.cell(row=current_row, column=3).alignment = left_align
    current_row += 1
    ws.cell(row=current_row, column=3, value="Kepala Sekolah").font = font_normal_12
    ws.cell(row=current_row, column=3).alignment = left_align

    tz_jakarta = pytz.timezone('Asia/Jakarta')
    now = datetime.now(tz_jakarta)

    # Format Tanggal: Titik-titik untuk Tanggal dan Bulan, Tahun sesuai tahun berlaku
    tanggal_formatted = f"Bantul, .................................... {now.year}"

    # Bantul, Guru Mapel di kolom K
    ws.cell(row=current_row - 1, column=11, value=tanggal_formatted).font = font_normal_12 # Column K is 11
    ws.cell(row=current_row - 1, column=11).alignment = left_align
    ws.cell(row=current_row, column=11, value=f"Guru {mapel}").font = font_normal_12
    ws.cell(row=current_row, column=11).alignment = left_align

    current_row += 3 # Jarak untuk tanda tangan (3 baris kosong)

    # Nama Kepala Sekolah dan Nama Guru
    ws.cell(row=current_row, column=3, value=KEPALA_SEKOLAH).font = font_normal_12
    ws.cell(row=current_row, column=3).alignment = left_align
    ws.cell(row=current_row, column=11, value=guru).font = font_normal_12 # Column K is 11
    ws.cell(row=current_row, column=11).alignment = left_align

    current_row += 1
    ws.cell(row=current_row, column=3, value=f"NIP. {NIP_KEPSEK}").font = font_normal_12
    ws.cell(row=current_row, column=3).alignment = left_align
    ws.cell(row=current_row, column=11, value=f"NIP. {nip_guru}").font = font_normal_12 # Column K is 11
    ws.cell(row=current_row, column=11).alignment = left_align

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ===================== APLIKASI STREAMLIT UTAMA =====================
st.set_page_config(layout="wide", page_title="Aplikasi Manajemen Data Siswa & Daftar Hadir")
st.title("Aplikasi Manajemen Data Siswa & Daftar Hadir")

default_csv_path = "daftar_siswa.csv"
df = None

# --- Memuat File CSV ---
# Menggunakan st.session_state untuk menjaga df setelah diunggah/dimuat
if 'df_loaded' not in st.session_state:
    st.session_state.df_loaded = None

if st.session_state.df_loaded is None:
    if os.path.exists(default_csv_path):
        try:
            df = pd.read_csv(default_csv_path)
            st.session_state.df_loaded = df
            st.success(f"File '{default_csv_path}' berhasil dimuat secara otomatis.")
        except Exception as e:
            st.error(f"Gagal memuat '{default_csv_path}': {e}. Coba unggah file manual.")

    if st.session_state.df_loaded is None: # Jika belum dimuat atau ada error
        st.warning(f"File '{default_csv_path}' tidak ditemukan atau gagal dimuat. Harap unggah file siswa.")
        uploaded_file_obj = st.file_uploader("Unggah file CSV/Excel daftar siswa (Harus punya kolom 'Kelas', 'Nama', 'NIS', 'Jenis_Kelamin')", type=["csv", "xlsx"], key="manual_upload")
        if uploaded_file_obj is not None:
            try:
                if uploaded_file_obj.name.endswith('.csv'):
                    df = pd.read_csv(uploaded_file_obj)
                else:
                    df = pd.read_excel(uploaded_file_obj)
                st.session_state.df_loaded = df
            except Exception as e:
                st.error(f"Terjadi kesalahan saat membaca file: {e}. Pastikan format file benar.")
                st.session_state.df_loaded = None
else:
    df = st.session_state.df_loaded # Ambil df dari session_state jika sudah dimuat

# Hentikan aplikasi jika DataFrame belum dimuat
if df is None:
    st.info("Silakan unggah file siswa untuk melanjutkan.")
    st.stop()

# Periksa kolom-kolom penting
required_cols_ds = ['Kelas', 'Nama', 'NIS', 'Jenis_Kelamin']
missing_cols_ds = [col for col in required_cols_ds if col not in df.columns]

if missing_cols_ds:
    st.error(f"Kolom yang diperlukan tidak ditemukan di file siswa Anda untuk 'Daftar Siswa': {', '.join(missing_cols_ds)}. Harap perbaiki file Anda.")
    st.stop()

st.markdown("---") # Garis pemisah


# Opsi kelas yang tersedia
kelas_options = sorted(df['Kelas'].unique())
if not kelas_options:
    st.error("Tidak ada data kelas yang ditemukan di kolom 'Kelas' file siswa Anda.")
    st.stop()

# --- Bagian 1: Cetak Daftar Siswa ---
st.header("📋 Cetak Daftar Siswa")

kelas_daftar_siswa = st.selectbox(
    "Pilih Kelas untuk Daftar Siswa",
    kelas_options,
    key="kelas_daftar_siswa_selectbox"
)
data_kelas_daftar_siswa = df[df['Kelas'] == kelas_daftar_siswa].reset_index(drop=True)

st.dataframe(data_kelas_daftar_siswa)

# Tambahkan informasi jumlah Laki-laki dan Perempuan di Streamlit
if 'Jenis_Kelamin' in data_kelas_daftar_siswa.columns:
    jumlah_L_selected_class = data_kelas_daftar_siswa[data_kelas_daftar_siswa['Jenis_Kelamin'].astype(str).str.upper() == 'L'].shape[0]
    jumlah_P_selected_class = data_kelas_daftar_siswa[data_kelas_daftar_siswa['Jenis_Kelamin'].astype(str).str.upper() == 'P'].shape[0]
    st.write(f"**Keterangan:**")
    st.write(f"- Jumlah Laki-laki (L) : {jumlah_L_selected_class}")
    st.write(f"- Jumlah Perempuan (P) : {jumlah_P_selected_class}")

nama_wali_kelas = st.text_input("Nama Wali Kelas", key="nama_wali_kelas_ds")
nip_wali_kelas = st.text_input("NIP Wali Kelas", key="nip_wali_kelas_ds")
semester_ds = st.selectbox(
    "Semester (Daftar Siswa)",
    ["Ganjil", "Genap"],
    key="semester_ds_selectbox"
)
#current_year = datetime.now().year
#tahun_options = [f"{y}/{y+1}" for y in range(current_year + 1, current_year - 5, -1)]
start_year = 2025
tahun_options = [f"{y}/{y+1}" for y in range(start_year, start_year + 5)] # Contoh: 2025/2026 hingga 2029/2030
tahun_ds = st.selectbox(
    "Tahun Pelajaran (Daftar Siswa)",
    tahun_options,
    index=0,
    key="tahun_ds_selectbox"
)

# Mengubah tombol generate dan download untuk Excel
if st.button("🔽 Generate Daftar Siswa (Excel)", key="generate_ds_button"):
    if not nama_wali_kelas or not nip_wali_kelas:
        st.error("Nama Wali Kelas dan NIP Wali Kelas tidak boleh kosong.")
    else:
        excel_buffer_ds = generate_excel_daftar_siswa(data_kelas_daftar_siswa, kelas_daftar_siswa, semester_ds, tahun_ds, nama_wali_kelas, nip_wali_kelas)
        st.download_button(
            label="Unduh Daftar Siswa (Excel)",
            data=excel_buffer_ds,
            file_name=f"Daftar_Siswa_{kelas_daftar_siswa}_{tahun_ds}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_ds_button"
        )
        st.success("Daftar Siswa berhasil dibuat!")

st.markdown("---")

# --- Bagian 2: Cetak Daftar Hadir Siswa ---
st.header("✍️ Cetak Daftar Hadir Siswa")

kelas_absensi = st.selectbox(
    "Pilih Kelas untuk Daftar Hadir",
    kelas_options,
    key="kelas_absensi_selectbox"
)
data_kelas_absensi = df[df['Kelas'] == kelas_absensi].reset_index(drop=True)

# Menggunakan selectbox untuk Mata Pelajaran
mapel_absensi = st.selectbox("Mata Pelajaran (Daftar Hadir)", COMMON_SUBJECTS, key="mapel_absensi_selectbox")

guru_absensi = st.text_input("Nama Guru (Daftar Hadir)", key="guru_absensi")
nip_guru_absensi = st.text_input("NIP Guru (Daftar Hadir)", key="nip_guru_absensi")
semester_absensi = st.selectbox(
    "Semester (Daftar Hadir)",
    ["Ganjil", "Genap"],
    key="semester_absensi_selectbox"
)
tahun_absensi = st.selectbox(
    "Tahun Pelajaran (Daftar Hadir)",
    tahun_options,
    index=0,
    key="tahun_absensi_selectbox"
)

# Mengubah tombol generate dan download untuk Excel
if st.button("🔽 Generate Daftar Hadir (Excel)", key="generate_absensi_button"):
    if not mapel_absensi or not guru_absensi or not nip_guru_absensi:
        st.error("Mata Pelajaran, Nama Guru, dan NIP Guru tidak boleh kosong.")
    else:
        excel_buffer_absensi = generate_excel_absensi_panjang(data_kelas_absensi, mapel_absensi, semester_absensi, kelas_absensi, tahun_absensi, guru_absensi, nip_guru_absensi)
        st.download_button(
            label="Unduh Daftar Hadir (Excel)",
            data=excel_buffer_absensi,
            file_name=f"Daftar_Hadir_{kelas_absensi}_{mapel_absensi}_{tahun_absensi}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_absensi_button"
        )
        st.success("Daftar Hadir berhasil dibuat!")

st.markdown("---")

# --- Bagian 3: Cetak Form Nilai Siswa ---
st.header("📊 Cetak Form Nilai Siswa")

kelas_nilai = st.selectbox(
    "Pilih Kelas untuk Form Nilai",
    kelas_options,
    key="kelas_nilai_selectbox"
)
data_kelas_nilai = df[df['Kelas'] == kelas_nilai].reset_index(drop=True)

mapel_nilai = st.selectbox("Mata Pelajaran (Form Nilai)", COMMON_SUBJECTS, key="mapel_nilai_selectbox")

guru_nilai = st.text_input("Nama Guru (Form Nilai)", key="guru_nilai")
nip_guru_nilai = st.text_input("NIP Guru (Form Nilai)", key="nip_guru_nilai")
semester_nilai = st.selectbox(
    "Semester (Form Nilai)",
    ["Ganjil", "Genap"],
    key="semester_nilai_selectbox"
)
tahun_nilai = st.selectbox(
    "Tahun Pelajaran (Form Nilai)",
    tahun_options,
    index=0,
    key="tahun_nilai_selectbox"
)

if st.button("🔽 Generate Form Nilai (Excel)", key="generate_nilai_button"):
    if not mapel_nilai or not guru_nilai or not nip_guru_nilai:
        st.error("Mata Pelajaran, Nama Guru, dan NIP Guru tidak boleh kosong.")
    else:
        excel_buffer_nilai = generate_excel_form_nilai_siswa(data_kelas_nilai, mapel_nilai, semester_nilai, kelas_nilai, tahun_nilai, guru_nilai, nip_guru_nilai)
        st.download_button(
            label="Unduh Form Nilai (Excel)",
            data=excel_buffer_nilai,
            file_name=f"Form_Nilai_{kelas_nilai}_{mapel_nilai}_{tahun_nilai}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_nilai_button"
        )
        st.success("Form Nilai berhasil dibuat!")

st.markdown("---")
st.write(f"Created by Mersi")
